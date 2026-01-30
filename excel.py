import io
import re
from decimal import Decimal
from typing import List, Dict, Any, Tuple
import openpyxl

# 업로드 템플릿(대표님 제공 파일) 헤더 매핑
# A:I = 항목, 구분, 모델명, 규격, 설계가, 수의계약가, 납품가, 사무실 재고, 비고
EXPECTED_HEADERS = ["항목","구분","모델명","규격","설계가","수의계약가","납품가","사무실 재고","비고"]

def _norm_header(v) -> str:
    return str(v).strip() if v is not None else ""

def parse_money(value) -> int:
    """'5,840,000원' 같은 문자열을 정수(원)로 변환."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    s = re.sub(r"[^0-9]", "", s)  # 숫자만 남김
    return int(s) if s else 0

def parse_int(value) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    s = re.sub(r"[^0-9-]", "", s)
    return int(s) if s else 0

def read_products_from_excel(file_bytes: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    """엑셀 업로드 파일(bytes)에서 제품 리스트를 추출."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    headers = [_norm_header(ws.cell(1, c).value) for c in range(1, 10)]
    errors = []
    if headers[:9] != EXPECTED_HEADERS:
        errors.append(f"엑셀 헤더가 템플릿과 다릅니다. 기대={EXPECTED_HEADERS}, 실제={headers[:9]}")

    items: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 10)]
        if all(v is None or str(v).strip()=="" for v in row):
            continue

        item_name = str(row[0]).strip() if row[0] is not None else ""
        category_name = str(row[1]).strip() if row[1] is not None else ""
        product_name = str(row[2]).strip() if row[2] is not None else ""
        spec = str(row[3]).strip() if row[3] is not None else ""

        price_design = parse_money(row[4])
        price_small = parse_money(row[5])  # 템플릿의 '수의계약가' = 소보수가
        price_delivery = parse_money(row[6])
        stock_qty = parse_int(row[7])
        memo = str(row[8]).strip() if row[8] is not None else ""

        # 최소 유효성: 모델명(제품명) 필수
        if not product_name:
            errors.append(f"{r}행: 모델명이 비어있습니다.")
            continue

        items.append({
            "item_name": item_name,
            "category_name": category_name,
            "name": product_name,
            "spec": spec,
            "price_design": price_design,
            "price_small": price_small,
            "price_delivery": price_delivery,
            "stock_qty": stock_qty,
            "memo": memo,
        })
    return items, errors

def write_products_to_excel(rows: List[Dict[str, Any]]) -> bytes:
    """DB 제품 목록을 템플릿 형식으로 엑셀 bytes로 생성."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # header
    for c, h in enumerate(EXPECTED_HEADERS, start=1):
        ws.cell(1, c).value = h

    for i, row in enumerate(rows, start=2):
        ws.cell(i, 1).value = row.get("item_name","")
        ws.cell(i, 2).value = row.get("category_name","")
        ws.cell(i, 3).value = row.get("name","")
        ws.cell(i, 4).value = row.get("spec","")
        ws.cell(i, 5).value = int(row.get("price_design",0) or 0)
        ws.cell(i, 6).value = int(row.get("price_small",0) or 0)
        ws.cell(i, 7).value = int(row.get("price_delivery",0) or 0)
        ws.cell(i, 8).value = int(row.get("stock_qty",0) or 0)
        ws.cell(i, 9).value = row.get("memo","")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
