from __future__ import annotations

from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from typing import Optional, Any, Dict, List
import io
import datetime

import openpyxl

from app.core.db import get_db
from app.core.deps import get_current_user
from app.models.user import User
from app.core.config import settings

# =========================================================
# 업링크 제품(자재관리) Router (DB 저장형 - v2.0)
#
# 정책:
# - 조회(GET /api/products): 로그인 없이도 가능(기존 구조 유지)
# - 등록/수정/삭제/업로드/다운로드: 관리자(6) 또는 운영자(7)만 허용
#
# ⚠️ 매우 중요:
# - /upload, /download 같은 "고정 경로"는 /{product_id} 보다 먼저 등록해야
#   Starlette 라우팅이 /upload를 {product_id}로 오인하지 않습니다.
# =========================================================

router = APIRouter(tags=["products"])

ROLE_ADMIN_ID = 6
ROLE_OPERATOR_ID = 7

EXPECTED_HEADERS = ["항목", "구분", "모델명", "규격", "설계가", "수의계약가", "납품가", "사무실 재고", "비고"]

def _role_id(user: Any) -> Optional[int]:
    rid = getattr(user, "role_id", None)
    try:
        return int(rid) if rid is not None else None
    except Exception:
        return None

def require_admin_or_operator(user: User = Depends(get_current_user)) -> User:
    rid = _role_id(user)
    if rid in (ROLE_ADMIN_ID, ROLE_OPERATOR_ID):
        return user
    raise HTTPException(status_code=403, detail=f"관리자/운영자만 가능합니다. (role_id={rid})")

def _has_column(db, col: str) -> bool:
    row = db.execute(
        text("""
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema = :schema AND table_name = 'products' AND column_name = :col
            LIMIT 1
        """),
        {"schema": settings.DB_SCHEMA, "col": col},
    ).fetchone()
    return bool(row)

def _coerce_num(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    # 콤마 제거
    s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return 0.0

def _coerce_int(v) -> int:
    if v is None:
        return 0
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = str(v).strip().replace(",", "")
    try:
        return int(float(s))
    except Exception:
        return 0

# =========================================================
# ✅ 업로드/다운로드는 반드시 /{product_id} 보다 먼저!
# =========================================================

@router.post("/upload")
def upload_products_xlsx(
    file: UploadFile = File(...),
    db=Depends(get_db),
    _user: User = Depends(require_admin_or_operator),
):
    """엑셀 업로드 -> products 테이블 UPSERT"""
    try:
        content = file.file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active

        header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        header = header[: len(EXPECTED_HEADERS)]
        if header != EXPECTED_HEADERS:
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "엑셀 헤더가 기준과 다릅니다.",
                    "expected": EXPECTED_HEADERS,
                    "got": header,
                },
            )

        # DB 컬럼 존재 여부에 따라 동적 적용
        has_item_name = _has_column(db, "item_name")
        has_category_name = _has_column(db, "category_name")
        has_unit = _has_column(db, "unit")
        has_is_active = _has_column(db, "is_active")

        upsert_cols = ["id", "name", "spec", "price_design", "price_small", "price_delivery", "stock_qty", "memo"]
        if has_item_name:
            upsert_cols.insert(1, "item_name")
        if has_category_name:
            # item_name 다음 위치
            idx = 2 if has_item_name else 1
            upsert_cols.insert(idx, "category_name")
        if has_unit:
            upsert_cols.append("unit")
        if has_is_active:
            upsert_cols.append("is_active")

        col_sql = ", ".join(upsert_cols)
        val_sql = ", ".join([f":{c}" for c in upsert_cols])

        update_cols = [c for c in upsert_cols if c != "id"]
        update_sql = ", ".join([f"{c}=EXCLUDED.{c}" for c in update_cols])

        sql = text(
            f"""INSERT INTO products ({col_sql})
            VALUES ({val_sql})
            ON CONFLICT (id) DO UPDATE SET {update_sql}
            """
        )

        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None:
                continue
            # EXPECTED_HEADERS 매핑:
            # 0 항목(item_name), 1 구분(category_name), 2 모델명(name), 3 규격(spec),
            # 4 설계가, 5 수의계약가(price_small), 6 납품가(price_delivery), 7 재고, 8 비고(memo)
            item_name = (str(r[0]).strip() if r[0] is not None else "")
            category_name = (str(r[1]).strip() if r[1] is not None else "")
            name = (str(r[2]).strip() if r[2] is not None else "")
            spec = (str(r[3]).strip() if r[3] is not None else "")
            price_design = _coerce_num(r[4])
            price_small = _coerce_num(r[5])
            price_delivery = _coerce_num(r[6])
            stock_qty = _coerce_num(r[7])
            memo = (str(r[8]).strip() if r[8] is not None else "")

            if not name:
                continue

            # id는 업로드 시 자동 증가 대신, DB max(id)+1로 부여 (엑셀에 id 컬럼이 없음)
            rows.append(
                {
                    "item_name": item_name,
                    "category_name": category_name,
                    "name": name,
                    "spec": spec,
                    "price_design": price_design,
                    "price_small": price_small,
                    "price_delivery": price_delivery,
                    "stock_qty": stock_qty,
                    "memo": memo,
                }
            )

        if not rows:
            return {"ok": True, "imported": 0, "inserted": 0, "updated": 0}

        # 현재 max(id)
        max_id = db.execute(text("SELECT COALESCE(MAX(id), 0) FROM products")).scalar() or 0

        inserted = 0
        updated = 0

        # name을 기준으로 기존 제품이 있으면 업데이트, 없으면 신규 insert (프론트/기존 운영 방식 유지)
        for row in rows:
            existing = db.execute(
                text("SELECT id FROM products WHERE name=:name LIMIT 1"),
                {"name": row["name"]},
            ).fetchone()

            if existing:
                pid = int(existing[0])
                updated += 1
            else:
                max_id += 1
                pid = int(max_id)
                inserted += 1

            params = {"id": pid, "name": row["name"], "spec": row["spec"], "price_design": row["price_design"],
                      "price_small": row["price_small"], "price_delivery": row["price_delivery"],
                      "stock_qty": row["stock_qty"], "memo": row["memo"]}

            if has_item_name:
                params["item_name"] = row["item_name"]
            if has_category_name:
                params["category_name"] = row["category_name"]
            if has_unit:
                # 업로드 파일엔 unit 컬럼이 없어서 기본값
                params["unit"] = "EA"
            if has_is_active:
                params["is_active"] = True

            db.execute(sql, params)

        db.commit()
        imported = inserted + updated
        return {"ok": True, "imported": imported, "inserted": inserted, "updated": updated}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"업로드 실패: {str(e)}")

@router.get("/download")
def download_products_xlsx(
    db=Depends(get_db),
    _user: User = Depends(require_admin_or_operator),
):
    """DB -> 엑셀 다운로드"""
    has_item_name = _has_column(db, "item_name")
    has_category_name = _has_column(db, "category_name")
    has_unit = _has_column(db, "unit")

    cols = ["name", "spec", "price_design", "price_small", "price_delivery", "stock_qty", "memo"]
    if has_item_name:
        cols.insert(0, "item_name")
    if has_category_name:
        cols.insert(1 if has_item_name else 0, "category_name")
    if has_unit:
        cols.append("unit")

    select_cols = ", ".join(cols)
    where = ""
    if _has_column(db, "is_active"):
        where = "WHERE is_active = true"

    data = db.execute(text(f"SELECT {select_cols} FROM products {where} ORDER BY item_name ASC, category_name ASC, name ASC, spec ASC, id ASC")).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append(EXPECTED_HEADERS)

    for r in data:
        row = dict(zip(cols, r))
        ws.append([
            row.get("item_name", ""),
            row.get("category_name", ""),
            row.get("name", ""),
            row.get("spec", ""),
            float(row.get("price_design", 0) or 0),
            float(row.get("price_small", 0) or 0),
            float(row.get("price_delivery", 0) or 0),
            float(row.get("stock_qty", 0) or 0),
            row.get("memo", ""),
        ])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"products_{ts}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

# =========================================================
# 조회 (공개)
# =========================================================

@router.get("")
def list_products(
    q: str = "",
    limit: int = 500,
    include_inactive: int = 0,
    db=Depends(get_db),
):
    """제품 목록 조회(검색)"""
    q = (q or "").strip()
    params: Dict[str, Any] = {"limit": min(max(limit, 1), 2000)}
    where_parts = []

    # soft delete 컬럼이 있으면 기본적으로 활성 제품만 노출
    has_is_active = _has_column(db, "is_active")
    if has_is_active and not include_inactive:
        where_parts.append("is_active = true")

    if q:
        where_parts.append("(name ILIKE :q OR spec ILIKE :q OR memo ILIKE :q OR item_name ILIKE :q OR category_name ILIKE :q)")
        params["q"] = f"%{q}%"

    where = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""

    rows = db.execute(
        text(
            f"""
            SELECT id,
                   COALESCE(item_name,'') AS item_name,
                   COALESCE(category_name,'') AS category_name,
                   COALESCE(name,'') AS name,
                   COALESCE(spec,'') AS spec,
                   COALESCE(price_design,0) AS price_design,
                   COALESCE(price_small,0) AS price_small,
                   COALESCE(price_delivery,0) AS price_delivery,
                   COALESCE(stock_qty,0) AS stock_qty,
                   COALESCE(memo,'') AS memo
            FROM products
            {where}
            ORDER BY id ASC
            LIMIT :limit
            """
        ),
        params,
    ).fetchall()

    return [
        {
            "id": int(r[0]),
            "item_name": r[1],
            "category_name": r[2],
            "name": r[3],
            "spec": r[4],
            "price_design": float(r[5] or 0),
            "price_small": float(r[6] or 0),
            "price_delivery": float(r[7] or 0),
            "stock_qty": float(r[8] or 0),
            "memo": r[9],
        }
        for r in rows
    ]

# =========================================================
# 등록/수정/삭제 (관리자/운영자)
# =========================================================

@router.post("")
def create_product(
    payload: Dict[str, Any],
    db=Depends(get_db),
    _user: User = Depends(require_admin_or_operator),
):
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="제품명(name)은 필수입니다.")

    # 새 id 부여
    new_id = int(db.execute(text("SELECT COALESCE(MAX(id),0)+1 FROM products")).scalar() or 1)

    cols = ["id", "item_name", "category_name", "name", "spec", "price_design", "price_small", "price_delivery", "stock_qty", "memo"]
    params = {
        "id": new_id,
        "item_name": (payload.get("item_name") or "").strip(),
        "category_name": (payload.get("category_name") or "").strip(),
        "name": name,
        "spec": (payload.get("spec") or "").strip(),
        "price_design": _coerce_num(payload.get("price_design")),
        "price_small": _coerce_num(payload.get("price_small")),
        "price_delivery": _coerce_num(payload.get("price_delivery")),
        "stock_qty": _coerce_num(payload.get("stock_qty")),
        "memo": (payload.get("memo") or "").strip(),
    }

    if _has_column(db, "unit"):
        cols.append("unit")
        params["unit"] = (payload.get("unit") or "EA").strip() or "EA"
    if _has_column(db, "is_active"):
        cols.append("is_active")
        params["is_active"] = True

    col_sql = ", ".join(cols)
    val_sql = ", ".join([f":{c}" for c in cols])

    db.execute(text(f"INSERT INTO products ({col_sql}) VALUES ({val_sql})"), params)
    db.commit()
    return {"id": new_id, **payload}

# ⚠️ PUT/PATCH/POST 모두 수정으로 허용 (프론트 호출 방식 차이 흡수)
@router.put("/{product_id}")
@router.patch("/{product_id}")
@router.post("/{product_id}")
def update_product(
    product_id: int,
    payload: Dict[str, Any],
    db=Depends(get_db),
    _user: User = Depends(require_admin_or_operator),
):
    name = (payload.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="제품명(name)은 필수입니다.")

    set_cols = [
        "item_name=:item_name",
        "category_name=:category_name",
        "name=:name",
        "spec=:spec",
        "price_design=:price_design",
        "price_small=:price_small",
        "price_delivery=:price_delivery",
        "stock_qty=:stock_qty",
        "memo=:memo",
    ]
    params = {
        "id": product_id,
        "item_name": (payload.get("item_name") or "").strip(),
        "category_name": (payload.get("category_name") or "").strip(),
        "name": name,
        "spec": (payload.get("spec") or "").strip(),
        "price_design": _coerce_num(payload.get("price_design")),
        "price_small": _coerce_num(payload.get("price_small")),
        "price_delivery": _coerce_num(payload.get("price_delivery")),
        "stock_qty": _coerce_num(payload.get("stock_qty")),
        "memo": (payload.get("memo") or "").strip(),
    }

    if _has_column(db, "unit") and "unit" in payload:
        set_cols.append("unit=:unit")
        params["unit"] = (payload.get("unit") or "EA").strip() or "EA"

    sql = text(f"UPDATE products SET {', '.join(set_cols)} WHERE id=:id")
    res = db.execute(sql, params)
    if res.rowcount == 0:
        db.rollback()
        raise HTTPException(status_code=404, detail="제품을 찾을 수 없습니다.")
    db.commit()
    return {"id": product_id, **payload}

@router.delete("/{product_id}")
def delete_product(
    product_id: int,
    db=Depends(get_db),
    _user: User = Depends(require_admin_or_operator),
):
    # soft delete 우선
    if _has_column(db, "is_active"):
        res = db.execute(text("UPDATE products SET is_active=false WHERE id=:id"), {"id": product_id})
        if res.rowcount == 0:
            db.rollback()
            raise HTTPException(status_code=404, detail="제품을 찾을 수 없습니다.")
        db.commit()
        return {"ok": True, "deleted": product_id, "mode": "soft"}

    # hard delete (FK로 막히면 409)
    try:
        res = db.execute(text("DELETE FROM products WHERE id=:id"), {"id": product_id})
        if res.rowcount == 0:
            db.rollback()
            raise HTTPException(status_code=404, detail="제품을 찾을 수 없습니다.")
        db.commit()
        return {"ok": True, "deleted": product_id, "mode": "hard"}
    except Exception:
        db.rollback()
        raise HTTPException(status_code=409, detail="참조 중인 제품이라 삭제할 수 없습니다. (비활성화 방식 권장)")
